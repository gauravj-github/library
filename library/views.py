from django.shortcuts import render, redirect
from django.views import View
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook

from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm


class AuthorCreateView(View):
    def get(self, request):
        form = AuthorForm()
        return render(request, 'library/author_form.html', {'form': form})

    def post(self, request):
        form = AuthorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('author-list')
        return render(request, 'library/author_form.html', {'form': form})


class BookCreateView(View):
    def get(self, request):
        form = BookForm()
        return render(request, 'library/book_form.html', {'form': form})

    def post(self, request):
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('book-list')
        return render(request, 'library/book_form.html', {'form': form})


class BorrowRecordCreateView(View):
    def get(self, request):
        form = BorrowRecordForm()
        return render(request, 'library/borrow_form.html', {'form': form})

    def post(self, request):
        form = BorrowRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('borrow-list')
        return render(request, 'library/borrow_form.html', {'form': form})

class AuthorListView(View):
    def get(self, request):
        authors = Author.objects.all()
        paginator = Paginator(authors, 5)
        page = request.GET.get('page')
        page_obj = paginator.get_page(page)
        return render(request, 'library/author_list.html', {'authors': page_obj})


class BookListView(View):
    def get(self, request):
        books = Book.objects.all()
        paginator = Paginator(books, 5)
        page = request.GET.get('page')
        page_obj = paginator.get_page(page)
        return render(request, 'library/book_list.html', {'books': page_obj})


class BorrowRecordListView(View):
    def get(self, request):
        borrows = BorrowRecord.objects.all()
        paginator = Paginator(borrows, 5)
        page = request.GET.get('page')
        page_obj = paginator.get_page(page)
        return render(request, 'library/borrow_list.html', {'borrows': page_obj})

class ExportExcelView(View):
    def get(self, request):
        wb = Workbook()

        # Authors Sheet
        ws1 = wb.active
        ws1.title = "Authors"
        ws1.append(["ID", "Name", "Email", "Bio"])
        for author in Author.objects.all():
            ws1.append([author.id, author.name, author.email, author.bio])

        # Books Sheet
        ws2 = wb.create_sheet("Books")
        ws2.append(["ID", "Title", "Genre", "Published Date", "Author"])
        for book in Book.objects.all():
            ws2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

        # Borrow Records Sheet
        ws3 = wb.create_sheet("Borrow Records")
        ws3.append(["ID", "User Name", "Book", "Borrow Date", "Return Date"])
        for record in BorrowRecord.objects.all():
            ws3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

        # Return Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb.save(response)
        return response

class HomeView(View):
    def get(self, request):
        return render(request, 'library/home.html')
